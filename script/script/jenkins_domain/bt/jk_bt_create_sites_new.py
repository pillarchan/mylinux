#!/usr/bin/env python3
#############################
## from demo.py
## requirement:
## python3 -m pip install requests
## python3 -m pip install openpyxl

import time,hashlib,sys,os,json,datetime
import logging
import re
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

class bt_api:
    __BT_KEY = 'CFDzCcYnkyLHouQu0pwfYgDB5ee9LwHB'
    __BT_PANEL = 'http://103.37.0.130:12580'
    __BT_NAME = "103.37.0.130"

    #如果希望多台面板，可以在实例化对象时，将面板地址与密钥传入
    def __init__(self,bt_panel = None,bt_key = None):
        if bt_panel:
            self.__BT_PANEL = bt_panel
            self.__BT_KEY = bt_key
        if (self.__BT_PANEL == 'http://23.248.251.242:8888'):
            self.__BT_NAME = '23.248.251.242'

    #计算MD5
    def __get_md5(self,s):
        m = hashlib.md5()
        m.update(s.encode('utf-8'))
        return m.hexdigest()

    #构造带有签名的关联数组
    def __get_key_data(self):
        now_time = int(time.time())
        p_data = {
                    'request_token':self.__get_md5(str(now_time) + '' + self.__get_md5(self.__BT_KEY)),
                    'request_time':now_time
                 }
        return p_data

    #发送POST请求并保存Cookie
    #@url 被请求的URL地址(必需)
    #@data POST参数，可以是字符串或字典(必需)
    #@timeout 超时时间默认1800秒
    #return string
    def __http_post_cookie(self,url,p_data,timeout=1800):
        cookie_file = '/tmp/' + self.__get_md5(self.__BT_PANEL) + '.cookie';
        if sys.version_info[0] == 2:
            #Python2
            import urllib,urllib2,ssl,cookielib

            #创建cookie对象
            cookie_obj = cookielib.MozillaCookieJar(cookie_file)

            #加载已保存的cookie
            if os.path.exists(cookie_file):cookie_obj.load(cookie_file,ignore_discard=True,ignore_expires=True)

            ssl._create_default_https_context = ssl._create_unverified_context

            data = urllib.urlencode(p_data)
            req = urllib2.Request(url, data)
            opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cookie_obj))
            response = opener.open(req,timeout=timeout)

            #保存cookie
            cookie_obj.save(ignore_discard=True, ignore_expires=True)
            return response.read()
        else:
            #Python3
            import urllib.request,ssl,http.cookiejar
            cookie_obj = http.cookiejar.MozillaCookieJar(cookie_file)
            if os.path.isfile(cookie_file):
                cookie_obj.load(cookie_file,ignore_discard=True,ignore_expires=True)
            handler = urllib.request.HTTPCookieProcessor(cookie_obj)
            data = urllib.parse.urlencode(p_data).encode('utf-8')
            req = urllib.request.Request(url, data)
            opener = urllib.request.build_opener(handler)
            response = opener.open(req,timeout = timeout)
            cookie_obj.save(ignore_discard=True, ignore_expires=True)
            result = response.read()
            if type(result) == bytes: result = result.decode('utf-8')
            return result

    def search_sites(self, search_str):
        url = self.__BT_PANEL + '/data?action=getData'

        p_data = self.__get_key_data()
        p_data['limit'] = 1000
        p_data['table'] = 'sites'
        #p_data['tojs'] = 'get_site_list'
        p_data['search'] = search_str

        result = self.__http_post_cookie(url,p_data)
        return json.loads(result)

    def search(self, sitelist):

        for i in sitelist:
            now = datetime.datetime.now()
            str_result = ''
            if (i != None):
                site = str(i).strip()
                search_rslt = self.search_sites(site)
                if (bool(search_rslt['data'])):
                    rslt_site = search_rslt['data'][0]['name']
                    rslt_siteId = search_rslt['data'][0]['id']
                    # print(search_rslt)
                    str_result = "%s\t%s" %(rslt_site,rslt_siteId)
                    print(str_result)
                else:
                    str_result = site
                    print(str_result)


    def create_site(self, site):
        """create a single site"""
        url = self.__BT_PANEL + '/site?action=AddSite'
        p_data = self.__get_key_data()

        webname = {}
        webname['domain'] = str(site).strip()
        webname['domainlist'] = ["www.%s" %webname['domain']]
        webname['count'] = 0
        path = '/www/wwwroot/%s' %site

        p_data['webname'] = json.dumps(webname)
        p_data['path'] = path
        p_data['type_id'] = 0
        p_data['type'] = '静态'
        p_data['version'] = ''
        p_data['port'] = 80
        p_data['ps'] = ''
        p_data['ftp'] = 'false'
        p_data['sql'] = 'false'

        # print(p_data)
        # jsondata = json.load(p_data)

        try:
            response = self.__http_post_cookie(url,p_data)
            # return response
            return json.loads(response)
        except Exception as ex:
            now = datetime.datetime.now()
            str_log = "error: %s %s %s\n" %(now.strftime("%Y/%m/%d, %H:%M:%S"),url,ex)
            print(str_log)

        # response = {"siteStatus": 'true', "siteId": 734, "ftpStatus": 'false', "databaseStatus": 'false'}


    def get_file(self,filepath):
        url = self.__BT_PANEL + '/files?action=GetFileBody'
        p_data = self.__get_key_data()
        p_data['path'] = filepath

        try:
            response = self.__http_post_cookie(url,p_data)
            print(response)
        except Exception as ex:
            now = datetime.datetime.now()
            str_log = "error: %s %s %s\n" %(now.strftime("%Y/%m/%d, %H:%M:%S"),url,ex)
            print(str_log)

    def save_file(self,filepath,content):
        url = url = self.__BT_PANEL + '/files?action=SaveFileBody'
        p_data = self.__get_key_data()

        p_data['path'] = filepath
        p_data['data'] = content
        p_data['encoding'] = 'utf-8'

        try:
            response = self.__http_post_cookie(url,p_data)
            return json.loads(response)
        except Exception as ex:
            now = datetime.datetime.now()
            str_log = "error: %s %s %s\n" %(now.strftime("%Y/%m/%d, %H:%M:%S"),url,ex)
            print(str_log)


    def create_sites(self,sitelist):
        """sitelist is a dict"""

        success_sites = {}
        for i in sitelist:
            str_result = {}
            site = str(i).strip()
            if (site != None):
                now = datetime.datetime.now()
                add_result = self.create_site(site)
                str_result[site] = {}
                if add_result.get('siteStatus') != None:
                    if (add_result['siteStatus'] == True):
                        str_result[site]['id'] = add_result['siteId']
                        success_sites[site] = add_result['siteId']
                        str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"),str_result)
                        print(str_log)

                        ###修改跳转URL
                        file = c_file()
                        data = file.set_index_html_new('/opt/script/jenkins_domain/bt/index_new.html',sitelist[i]['upm'],sitelist[i]['pid'],sitelist[i]['sid']) 
                        path = path = '/www/wwwroot/%s/index.html' %site
                        sav_result = self.save_file(path,data)
                        str_result[site] = "修改跳转URL: %s" %sav_result
                        str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                        print(str_log)
                    else:
                        str_result[site]['error'] = "%s" %add_result
                else:
                    str_result[site]['error'] = '%s' %add_result


        ## 申请证书
        if bool(success_sites):
            for site in success_sites:
                now = datetime.datetime.now()
                str_result = {}
                ssl_result = self.apply_cert(site,success_sites[site])

                if (ssl_result['status'] == True):
                    # print log
                    str_result[site] = "申请证书成功: %s" %ssl_result['status']
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                    print(str_log)

                    ## set ssl
                    time.sleep(1)
                    now = datetime.datetime.now()
                    key = ssl_result['private_key']
                    src  = ssl_result['cert'] + ssl_result['root']
                    set_ssl_rsp = self.set_ssl(site,key,src)
                    str_result[site] = "保存证书: %s" %set_ssl_rsp
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                    print(str_log)

                    ## http to https
                    time.sleep(1)
                    now = datetime.datetime.now()
                    set_https_rsp = self.set_to_https(site)
                    str_result[site] = "强制HTTPS: %s" %set_https_rsp
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                    print(str_log)

                    ## 修改 跳转URL
                    #time.sleep(2)
                    #now = datetime.datetime.now()
                    #file = c_file()
                    #data = file.set_index_html_new('/opt/script/jenkins_domain/bt/index_new.html',sitelist[i]['upm'],sitelist[i]['pid'],sitelist[i]['sid'])
                    #path = path = '/www/wwwroot/%s/index.html' %site
                    #sav_result = self.save_file(path,data)
                    #str_result[site] = "修改跳转URL: %s" %sav_result
                    #str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                    #print(str_log)
                else:
                    str_result[site] = "申请证书失败: %s" %ssl_result
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"),str_result)
                    print(str_log)


    def set_to_https(self,site):
        url = self.__BT_PANEL + '/site?action=HttpToHttps'
        p_data = self.__get_key_data()
        p_data['siteName'] = site

        try:
            response = self.__http_post_cookie(url,p_data)
            return json.loads(response)
        except Exception as ex:
            now = datetime.datetime.now()
            str_log = "error: %s %s %s\n" %(now.strftime("%Y/%m/%d, %H:%M:%S"),url,ex)
            print(str_log)

    ## 已创建的站点 修改index.html跳转连接
    def replace_index_html(self,sitelist):
        """sitelist is a dict {site: URL}"""
        now = datetime.datetime.now()

        for i in sitelist:
            str_result = {}
            if (i != None):
                site = str(i).strip()
                str_result[site] = {}
                ## 跳转后域名
                file = c_file()
                data = file.set_index_html_new('/opt/script/jenkins_domain/bt/index_new.html',sitelist[i]['upm'],sitelist[i]['pid'],sitelist[i]['sid'])
                ## 保存到bt
                path = '/www/wwwroot/%s/index.html' %site
                sav_result = self.save_file(path,data)

                #if '"status": true' in sav_result:
                #    str_result[site]['modified'] = "成功"
                #    str_result[site]['path'] = "%s %s" %(path,URL)
                #else:
                #    str_result[site]['modified'] = "失败"
                #    str_result[site]['error'] = "%s" %(sav_result)
                #    str_result[site]['debug'] = "%s %s" %(path,URL)
                str_result[site] = "%s" %(sav_result)

                str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"),str_result)
                # logging.info(str_log)
                print(str_log)

    def apply_cert(self,site,siteId):
        # test rted.cc
        # LetsEncrypt
        url = self.__BT_PANEL + '/acme?action=apply_cert_api'
        www = "www.%s" %site
        domains = []
        domains.append(site)
        domains.append(www)

        p_data = self.__get_key_data()
        p_data['domains'] = json.dumps(domains)
        p_data['auth_type'] = "http"
        p_data['auth_to'] = siteId
        p_data['auto_wildcard'] = 0
        p_data['id'] = siteId
        # print(p_data)

        try:
            response = self.__http_post_cookie(url,p_data)
            return json.loads(response)
        except Exception as ex:
            now = datetime.datetime.now()
            str_log = "error: %s %s %s\n" %(now.strftime("%Y/%m/%d, %H:%M:%S"),url,ex)
            print(str_log)
            pass

    def only_apply_cert(self,sitelist):
        """sitelist is a dict {site: siteId}"""

        for i in sitelist:
            now = datetime.datetime.now()
            str_result = {}
            site = str(i).strip()
            if (site != None):

                ssl_result = self.apply_cert(site,sitelist[i])
                str_result[site] = ssl_result

                if (ssl_result['status'] == True):
                    # print log
                    str_result[site] = "申请证书成功: %s" %ssl_result['status']
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                    print(str_log)

                    ## set ssl
                    time.sleep(2)
                    now = datetime.datetime.now()
                    key = ssl_result['private_key']
                    src  = ssl_result['cert'] + ssl_result['root']
                    set_ssl_rsp = self.set_ssl(site,key,src)
                    str_result[site] = "保存证书: %s" %set_ssl_rsp
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                    print(str_log)

                    ## http to https
                    time.sleep(2)
                    now = datetime.datetime.now()
                    set_https_rsp = self.set_to_https(site)
                    str_result[site] = "强制HTTPS: %s" %set_https_rsp
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"), str_result)
                    print(str_log)
                else:
                    str_result[site] = "申请证书失败: %s" %ssl_result
                    str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"),str_result)
                    print(str_log)


    def set_ssl(self,site,key,csr):
        url = self.__BT_PANEL + '/site?action=SetSSL'

        p_data = self.__get_key_data()
        p_data['type'] = 1
        p_data['siteName'] = site
        p_data['key'] = key
        p_data['csr'] = csr

        try:
            response = self.__http_post_cookie(url,p_data)
            return json.loads(response)
        except Exception as ex:
            now = datetime.datetime.now()
            str_log = "error: %s %s %s\n" %(now.strftime("%Y/%m/%d, %H:%M:%S"),url,ex)
            print(str_log)

    def only_create_sites(self,sitelist):
        """sitelist is a dict {site: URL}"""
        now = datetime.datetime.now()

        success_sites = {}
        for i in sitelist:
            str_result = {}
            if (i != None):
                site = str(i).strip()
                add_result = self.create_site(site)
                str_result[site] = {}
                if add_result.get('siteStatus') != None:
                    if (add_result['siteStatus'] == True):
                        str_result[site]['success'] = add_result['siteStatus']
                        success_sites[site] = add_result['siteId']
                        ##跳转后域名
                        file = c_file()
                        data = file.set_index_html_new('/opt/script/jenkins_domain/bt/index_new.html',sitelist[i]['upm'],sitelist[i]['pid'],sitelist[i]['sid'])
                        ## 保存到bt
                        path = path = '/www/wwwroot/%s/index.html' %site
                        sav_result = self.save_file(path,data)

                        if '"status": true' in sav_result:
                            str_result[site]['path'] = "%s %s" %(path,URL)
                        else:
                            str_result[site]['path'] = "失败"
                            str_result[site]['error'] = "%s %s %s" %(sav_result,path,URL)
                    else:
                        str_result[site]['error'] = "%s" %add_result
                else:
                    str_result[site]['error'] = '%s' %add_result

                str_log = "%s %s" %(now.strftime("%Y/%m/%d, %H:%M:%S"),str_result)
                # logging.info(str_log)
                print(str_log)



class c_file:

    def read_xlsx(self,exlfile):
        """ read every row A and B values into a dict """

        self.wb = load_workbook(exlfile)
        ws = self.wb.worksheets[0]
        #delete title row
        ws.delete_rows(0)

        p_dict = {}
        for row in ws.values:
            p_dict[row[0]] = row[1]

        return p_dict

    def set_index_html(self,f_index,URL):

        f = open(f_index, 'r')
        html_content = f.read()
        f.close()

        str_repl = "window.location.href = '%s'" %str(URL).strip()
        replaced = re.sub('window\.location\.href.*',str_repl,html_content)

        return replaced

    def read_xlsx_new(self,exlfile):
        ## col A : sitename
        ## col B : upm
        ## col C : pid
        ## col D : sid

        self.wb = load_workbook(exlfile)
        ws = self.wb.worksheets[0]
        #delete title row
        ws.delete_rows(0)

        p_dict = {}
        for row in ws.values:
            site = str(row[0]).strip()
            if (site != None):
                p_dict[site] = {}
                p_dict[site]['upm'] = str(row[1]).strip()
                p_dict[site]['pid'] = str(row[2]).strip()
                p_dict[site]['sid'] = str(row[3]).strip()

        return p_dict


    def set_index_html_new(self,f_index,upm,pid,sid):
        ##https://${randomStr(12)}.${domain}:60443/jpm?upm=用户id&pid=推广路径&sid=1001
        ## read file index.html
        f = open(f_index, 'r')
        html_content = f.read()
        f.close()
        ## replace rewrite_url
        upm = str(upm).strip()
        pid = str(pid).strip()
        sid = str(sid).strip()
        str_repl="upm=%s&pid=%s&sid=%s`" %(upm,pid,sid)

        replaced = re.sub('upm.*',str_repl,html_content)

        return replaced


if __name__ == '__main__':
    script,filename,action = sys.argv
    file = c_file()
    sitelist = file.read_xlsx_new(filename)
    # print(action)
    str_action = str(action)

    bt_tz = bt_api()

    if str_action == 'create_sites':
        bt_tz.create_sites(sitelist)
    elif str_action == 'only_create_sites':
        bt_tz.only_create_sites(sitelist)
    elif  str_action == 'search':
        bt_tz.search(sitelist)
    elif  str_action == 'apply_cert':
        bt_tz.only_apply_cert(sitelist)
    elif  str_action == 'replace_URL':
        bt_tz.replace_index_html(sitelist)
    else:
        print('操作不存在')
